import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import urllib2
from django.utils.encoding import smart_str, smart_unicode
from bs4 import BeautifulSoup
from pathlib import Path
import sys

#fileName="GeeksforGeeks.xlsx"
#SheetName="Arrays"
#quote_page = "https://www.geeksforgeeks.org/array-data-structure/"

with open('./setting.properties') as fp: 
    quote_page=str(fp.readline())
    fileName=str(fp.readline())
    SheetName=str(fp.readline())
quote_page=quote_page.rstrip('\n')
fileName=fileName.rstrip('\n')
SheetName=SheetName.rstrip('\n')


page = urllib2.urlopen(quote_page)
soup = BeautifulSoup(page, 'html.parser')
li_list=soup.find_all('li')

pattern='"https://www.geeksforgeeks.org/"[A-Za-z]+'

count=1

my_file = Path(fileName)
if my_file.is_file():
    wb = load_workbook(filename = fileName)
    ws = wb.create_sheet(SheetName)
    ws.title = SheetName
#    print "file already exists"
#    sys.exit(0)
    
else:
    wb = Workbook()
    ws = wb.create_sheet(SheetName)
    ws.title = SheetName

for something in li_list:
    for anchor_tag in something.find_all('a',href=True):
        link=str(anchor_tag['href'])
        if "https://www.geeksforgeeks.org/" not in link:
            continue
        text=smart_str(anchor_tag.string)
        if text.count(' ')<=1:
            continue
        new_link="=HYPERLINK(\""+link+"\",\"link\")"
        ws.cell(row=count,column=1).value=text;
        ws.cell(row=count,column=3).value=link;
        ws.cell(row=count,column=2).value=new_link;
        count=count+1

wb.save(fileName)


